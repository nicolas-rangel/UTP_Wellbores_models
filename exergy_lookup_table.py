# %%
"""
Specific physical exergy lookup-table generator.

Builds an Excel workbook with:
    1) An "Exergy Calculator" sheet where the user enters temperature, pressure and
       mass flow rate, and reads back interpolated enthalpy, entropy, specific
       exergy, gross heat rate and exergy rate.
    2) Two lookup sheets (enthalpy and entropy) over a temperature-pressure grid,
       computed with CoolProp.

The calculator sheet interpolates from the lookup sheets using INDEX/MATCH, so the
workbook is self-contained and works without Python once generated.

Specific physical exergy relative to a dead state (T0, P0):
    ex = (h - h0) - T0 * (s - s0)
"""

import numpy as np
import CoolProp.CoolProp as CP
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
fluid = 'Water'
T0_C = 20.0          # dead-state temperature [°C]
P0_bar = 1.01325     # dead-state pressure [bar] (1 atm)
T0 = T0_C + 273.15   # [K]
P0 = P0_bar * 1e5    # [Pa] -- SI for CoolProp

# Temperature and pressure ranges.
# Temperatures: 1 to 200 °C (fine step for interpolation).
temperatures_C = [round(t, 1) for t in np.linspace(1, 200, 200)]
# Pressures: 1 to 200 bar.
pressures_bar = [round(p, 3) for p in np.linspace(1, 200, 200)]

# Dead-state properties (or the reinjection state, depending on the case).
h0 = CP.PropsSI('H', 'T', T0, 'P', P0, fluid)  # J/kg
s0 = CP.PropsSI('S', 'T', T0, 'P', P0, fluid)  # J/(kg·K)


# ---------------------------------------------------------------------------
# Data generation
# ---------------------------------------------------------------------------
def get_property(prop, T_C, P_bar, fluid='Water'):
    """Return the property in kJ/kg (enthalpy) or kJ/(kg·K) (entropy), or None if out of range."""
    T_K = T_C + 273.15
    P_Pa = P_bar * 1e5
    try:
        val = CP.PropsSI(prop, 'T', T_K, 'P', P_Pa, fluid)
        if prop == 'H':
            return val / 1000  # kJ/kg
        elif prop == 'S':
            return val / 1000  # kJ/(kg·K)
    except Exception:
        return None


h_data = {}  # (T_C, P_bar) -> h [kJ/kg]
s_data = {}  # (T_C, P_bar) -> s [kJ/(kg·K)]

for T in temperatures_C:
    for P in pressures_bar:
        h_data[(T, P)] = get_property('H', T, P, fluid)
        s_data[(T, P)] = get_property('S', T, P, fluid)

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
input_font = Font(name='Arial', size=10, color='0000FF')
input_fill = PatternFill('solid', fgColor='FFF2CC')
calc_font = Font(name='Arial', size=10)
title_font = Font(name='Arial', bold=True, size=12, color='2F5496')
subtitle_font = Font(name='Arial', bold=True, size=10, color='2F5496')
data_font = Font(name='Arial', size=9)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
center_align = Alignment(horizontal='center', vertical='center')

wb = Workbook()

nT = len(temperatures_C)
nP = len(pressures_bar)
last_col = get_column_letter(1 + nP)
h0_kJ = h0 / 1000
s0_kJ = s0 / 1000

# ---------------------------------------------------------------------------
# Sheet 1: Exergy calculator
# ---------------------------------------------------------------------------
ws_calc = wb.active
ws_calc.title = 'Exergy Calculator'
ws_calc.sheet_properties.tabColor = '2F5496'

ws_calc['A1'] = 'Specific Physical Exergy Calculator'
ws_calc['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
ws_calc.merge_cells('A1:I1')

ws_calc['A2'] = f'Working fluid: {fluid} | Dead state: T₀ = {T0_C} °C, P₀ = {P0_bar} bar'
ws_calc['A2'].font = Font(name='Arial', size=10, italic=True, color='808080')
ws_calc.merge_cells('A2:I2')

ws_calc['A3'] = f'Dead state: h₀ = {h0_kJ:.4f} kJ/kg, s₀ = {s0_kJ:.6f} kJ/(kg·K)'
ws_calc['A3'].font = Font(name='Arial', size=10, italic=True, color='808080')
ws_calc.merge_cells('A3:I3')

headers = [
    'Point', 'T [°C]', 'P [bar]', 'h [kJ/kg]', 's [kJ/(kg·K)]',
    'Exergy [kJ/kg]', 'ṁ [kg/s]',
    'Gross Heat Rate [kW]', 'Exergy Rate [kW]',
]
col_widths = [10, 12, 12, 16, 16, 16, 14, 22, 20]
for j, (hdr, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws_calc.cell(row=5, column=j, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    ws_calc.column_dimensions[get_column_letter(j)].width = w

n_rows = 20
for i in range(n_rows):
    row = 6 + i
    ws_calc.cell(row=row, column=1, value=i + 1).font = calc_font
    ws_calc.cell(row=row, column=1).alignment = center_align

    for col in [2, 3]:
        cell = ws_calc.cell(row=row, column=col)
        cell.font = input_font
        cell.fill = input_fill
        cell.alignment = center_align
        cell.number_format = '0.0'

    # References to the lookup sheets.
    T_ref = f"Enthalpy!$A$3:$A${2 + nT}"
    P_ref = f"Enthalpy!$B$2:${last_col}$2"
    data_ref_h = f"Enthalpy!$B$3:${last_col}${2 + nT}"
    T_ref_s = f"Entropy!$A$3:$A${2 + nT}"
    P_ref_s = f"Entropy!$B$2:${last_col}$2"
    data_ref_s = f"Entropy!$B$3:${last_col}${2 + nT}"

    T_cell = f"B{row}"
    P_cell = f"C{row}"

    # Enthalpy formula.
    h_formula = f'=IFERROR(INDEX({data_ref_h},MATCH({T_cell},{T_ref},1),MATCH({P_cell},{P_ref},1)),"—")'
    cell_h = ws_calc.cell(row=row, column=4, value=h_formula)
    cell_h.font = calc_font
    cell_h.alignment = center_align
    cell_h.number_format = '0.00'

    # Entropy formula.
    s_formula = f'=IFERROR(INDEX({data_ref_s},MATCH({T_cell},{T_ref_s},1),MATCH({P_cell},{P_ref_s},1)),"—")'
    cell_s = ws_calc.cell(row=row, column=5, value=s_formula)
    cell_s.font = calc_font
    cell_s.alignment = center_align
    cell_s.number_format = '0.0000'

    # Exergy formula: ex = (h - h0) - T0*(s - s0).
    exergy_formula = f'=IFERROR((D{row}-{h0_kJ:.6f})-{T0}*(E{row}-{s0_kJ:.8f}),"—")'
    cell_ex = ws_calc.cell(row=row, column=6, value=exergy_formula)
    cell_ex.font = Font(name='Arial', size=10, bold=True)
    cell_ex.alignment = center_align
    cell_ex.number_format = '0.00'

    # Column G: mass flow rate (user input).
    cell_mdot = ws_calc.cell(row=row, column=7)
    cell_mdot.font = input_font
    cell_mdot.fill = input_fill
    cell_mdot.alignment = center_align
    cell_mdot.number_format = '0.00'

    # Column H: gross heat rate relative to the dead state = m_dot * (h - h0) [kW].
    gross_heat_formula = f'=IFERROR(G{row}*(D{row}-{h0_kJ:.6f}),"—")'
    cell_gh = ws_calc.cell(row=row, column=8, value=gross_heat_formula)
    cell_gh.font = Font(name='Arial', size=10, bold=True)
    cell_gh.alignment = center_align
    cell_gh.number_format = '0.00'

    # Column I: exergy rate relative to the dead state = m_dot * ex [kW].
    exergy_rate_formula = f'=IFERROR(G{row}*F{row},"—")'
    cell_er = ws_calc.cell(row=row, column=9, value=exergy_rate_formula)
    cell_er.font = Font(name='Arial', size=10, bold=True)
    cell_er.alignment = center_align
    cell_er.number_format = '0.00'

    for col in range(1, 10):
        ws_calc.cell(row=row, column=col).border = thin_border

# Notes
note_row = 6 + n_rows + 1
ws_calc.cell(row=note_row, column=1, value='Notes:').font = subtitle_font
ws_calc.cell(row=note_row + 1, column=1,
             value='• Enter Temperature [°C], Pressure [bar], and Mass flow rate [kg/s] in the blue cells (columns B, C, G).').font = data_font
ws_calc.cell(row=note_row + 2, column=1,
             value='• h and s are interpolated from the lookup tables via INDEX/MATCH (nearest lower value).').font = data_font
ws_calc.cell(row=note_row + 3, column=1,
             value=f'• Exergy = (h − h₀) − T₀·(s − s₀), with T₀ = {T0} K, h₀ = {h0_kJ:.4f} kJ/kg, s₀ = {s0_kJ:.6f} kJ/(kg·K).').font = data_font
ws_calc.cell(row=note_row + 4, column=1,
             value='• Gross Heat Rate [kW] = ṁ × (h − h₀). Total thermal power relative to the dead state.').font = data_font
ws_calc.cell(row=note_row + 5, column=1,
             value='• Exergy Rate [kW] = ṁ × Exergy. Total exergetic potential relative to the dead state.').font = data_font
ws_calc.cell(row=note_row + 6, column=1,
             value='• "—" indicates the (T, P) pair is outside the lookup table range or in a two-phase region.').font = data_font


# ---------------------------------------------------------------------------
# Helper to write a lookup sheet
# ---------------------------------------------------------------------------
def write_lookup_sheet(ws, sheet_name, data_dict, value_label, num_fmt, tab_color):
    ws.title = sheet_name
    ws.sheet_properties.tabColor = tab_color

    ws['A1'] = f'{value_label} — {fluid}'
    ws['A1'].font = title_font
    ws.merge_cells(f'A1:{get_column_letter(1 + nP)}1')

    ws.cell(row=2, column=1, value='T [°C] \\ P [bar]').font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).alignment = center_align
    ws.column_dimensions['A'].width = 18

    for j, P in enumerate(pressures_bar, 2):
        cell = ws.cell(row=2, column=j, value=P)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(j)].width = 13

    for i, T in enumerate(temperatures_C):
        row = 3 + i
        cell_t = ws.cell(row=row, column=1, value=T)
        cell_t.font = Font(name='Arial', bold=True, size=9)
        cell_t.alignment = center_align
        cell_t.fill = PatternFill('solid', fgColor='D6E4F0')

        for j, P in enumerate(pressures_bar, 2):
            val = data_dict.get((T, P))
            cell = ws.cell(row=row, column=j)
            if val is not None:
                cell.value = round(val, 6)
                cell.number_format = num_fmt
            else:
                cell.value = '—'
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

    ws.freeze_panes = 'B3'


# ---------------------------------------------------------------------------
# Sheet 2: enthalpy lookup table
# ---------------------------------------------------------------------------
ws_h = wb.create_sheet()
write_lookup_sheet(ws_h, 'Enthalpy', h_data, 'Specific Enthalpy h [kJ/kg]', '0.00', '548235')

# ---------------------------------------------------------------------------
# Sheet 3: entropy lookup table
# ---------------------------------------------------------------------------
ws_s = wb.create_sheet()
write_lookup_sheet(ws_s, 'Entropy', s_data, 'Specific Entropy s [kJ/(kg·K)]', '0.000000', 'BF8F00')

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------
out_path = 'Exergy_Lookup_Table.xlsx'
wb.save(out_path)
print(f'Saved: {out_path}')
print(f'  Temperatures: {len(temperatures_C)} values ({temperatures_C[0]}–{temperatures_C[-1]} °C)')
print(f'  Pressures:    {len(pressures_bar)} values ({pressures_bar[0]}–{pressures_bar[-1]} bar)')
print(f'  Dead state: h0 = {h0_kJ:.4f} kJ/kg, s0 = {s0_kJ:.6f} kJ/(kg·K)')
